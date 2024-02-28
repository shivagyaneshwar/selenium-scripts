from openpyxl import Workbook
from win32com import client
import openpyxl
import mailchimp_marketing as MailchimpMarketing
from mailchimp_marketing.api_client import ApiClientError
import hashlib
import re




wb = openpyxl.load_workbook(r"C:\Users\shiva\Documents\Book1.xlsx")
sheet = wb["Sheet1"]
for i in range(2,sheet.max_row+1):
    # firstname = (sheet.cell(row=i,column=1).value)
    # lastname = (sheet.cell(row=i,column=2).value)
    email = (sheet.cell(row=i,column=3).value)
    try:
        client = MailchimpMarketing.Client()
        client.set_config({
        "api_key": "c0d3f7c35a59aaef9c9c27376b89ef59-us21","server": "us21"})
        response = client.lists.set_list_member("230842e0la", "subscriber_hash", {"email_address": email , "status_if_new": "subscribed"})
        email = re.sub('([A-Z]{1})', r'\1',email).lower()
        email = email.encode('utf-8')
        result = hashlib.md5(email).hexdigest()
        try:
            client = MailchimpMarketing.Client()
            client.set_config({
            "api_key": "c0d3f7c35a59aaef9c9c27376b89ef59-us21","server": "us21"})
            response = client.lists.update_list_member("230842e0la", result)
        except ApiClientError as error:
            print("Unuccessful In Updating"+ str(email))
    except ApiClientError as error:
        print("Unsuccessful In Adding - User Already Exists " + str(email))